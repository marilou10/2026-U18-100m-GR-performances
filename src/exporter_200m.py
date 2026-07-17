import os
import csv
import sys
import subprocess
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from config_200m import G
from normalizer import fmt_wind, fmt_comp, fmt_loc
try:
    from fpdf import FPDF
    _HAS_PDF = True
except ImportError:
    _HAS_PDF = False


WIND_FILL = PatternFill(start_color="FFFCE4", end_color="FFFCE4", fill_type="solid")


def make_output_dir():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, "output")
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def output_filename(output_dir):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    return os.path.join(output_dir, f"2026_U18_200m_GR_{timestamp}.xlsx")


def create_excel(all_results, ranking, wind_legal_ranking, wind_aided_ranking, fmt_note, fmt_wind, fmt_comp, fmt_loc):
    filename = output_filename(make_output_dir())
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "All_Performances"
    ws1.append(G)

    sorted_all = sorted(all_results, key=lambda x: float(x["performance"].replace("w","").replace("h","")) if x.get("performance","").replace("w","").replace("h","").replace(".","").isdigit() else 999)
    for i, r in enumerate(sorted_all, 1):
        ws1.append([i, r["name"], r["birth_year"], r["club"], r["performance"], fmt_wind(r["wind"]), fmt_comp(r), r["date"], fmt_loc(r), r["heat"], r["lane"], fmt_note(r)])

    ws2 = wb.create_sheet("Season_Best")
    ws2.append(["200 Μ ΚΟΡΑΣΙΔΩΝ (Κ18) 2026"])
    ws2.append(G)
    for i, r in enumerate(ranking, 1):
        ws2.append([i, r["name"], r["birth_year"], r["club"], r["performance"], fmt_wind(r["wind"]), fmt_comp(r), r["date"], fmt_loc(r), r["heat"], r["lane"], fmt_note(r)])

    ws3 = wb.create_sheet("Καλύτερες_Επιδόσεις")
    ws3.append(["200 Μ ΚΟΡΑΣΙΔΩΝ (Κ18) 2026"])
    ws3.append(G)

    for i, r in enumerate(wind_legal_ranking, 1):
        ws3.append([i, r["name"], r["birth_year"], r["club"], r["performance"], fmt_wind(r["wind"]), fmt_comp(r), r["date"], fmt_loc(r), r["heat"], r["lane"], fmt_note(r)])

    if wind_aided_ranking:
        ws3.append([])
        ws3.append(["ΜΕ ΑΝΕΜΟ"] + [""] * 11)
        for i, r in enumerate(wind_aided_ranking, 1):
            ws3.append([i, r["name"], r["birth_year"], r["club"], r["performance"], fmt_wind(r["wind"]), fmt_comp(r), r["date"], fmt_loc(r), r["heat"], r["lane"], fmt_note(r)])

    copyright_text = f"Copyright (c) {datetime.now().year} Μαρία Ελένη Αντωνοπούλου — Licensed under MIT"
    for ws in [ws1, ws2, ws3]:
        ws.append([])
        ws.append([copyright_text] + [""] * 11)

    for ws in [ws1, ws2, ws3]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            perf_cell = row[4]
            if perf_cell.value and isinstance(perf_cell.value, str) and "w" in perf_cell.value:
                for cell in row:
                    cell.fill = WIND_FILL

    wb.save(filename)
    return filename


def export_csv(all_results, filename, fmt_wind, fmt_comp, fmt_loc, fmt_note):
    csv_filename = filename.replace(".xlsx", ".csv")
    with open(csv_filename, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(G)
        for r in sorted(all_results, key=lambda x: float(x["performance"].replace("w","").replace("h","")) if x.get("performance","").replace("w","").replace("h","").replace(".","").isdigit() else 999):
            w.writerow([r["name"], r["birth_year"], r["club"], r["performance"],
                        fmt_wind(r["wind"]), fmt_comp(r), r["date"], fmt_loc(r),
                        r["heat"], r["lane"], fmt_note(r)])
    print(f"CSV file: {csv_filename}")


def export_pdf(all_results, wind_legal_ranking, wind_aided_ranking, filename, rget, fmt_wind, fmt_comp, fmt_loc, fmt_note):
    if not _HAS_PDF:
        return
    pdf_filename = filename.replace(".xlsx", ".pdf")

    copyright_text = f"Copyright (c) {datetime.now().year} Μαρία Ελένη Αντωνοπούλου — Licensed under MIT"

    class PDF(FPDF):
        def footer(self):
            self.set_y(-10)
            self.set_font("DejaVu", "", 5)
            self.cell(0, 5, copyright_text, align="C")

    pdf = PDF(orientation="L", unit="mm", format="A3")
    pdf.set_left_margin(5)

    _FONT_DIRS = [
        os.path.dirname(os.path.abspath(__file__)),
        "/usr/share/fonts/truetype/dejavu",
        "/usr/local/share/fonts/dejavu",
        os.path.expanduser("~/.fonts"),
    ]
    if sys.platform == "win32":
        _FONT_DIRS.insert(0, r"C:\Windows\Fonts")
    _TTF = None
    _TTF_B = None
    for d in _FONT_DIRS:
        r = os.path.join(d, "DejaVuSans.ttf")
        b = os.path.join(d, "DejaVuSans-Bold.ttf")
        if os.path.exists(r) and os.path.exists(b):
            _TTF, _TTF_B = r, b
            break
    if _TTF is None:
        print("[WARN] DejaVuSans.ttf not found — PDF may not render correctly")
    else:
        pdf.add_font("DejaVu", "", _TTF)
        pdf.add_font("DejaVu", "B", _TTF_B)

    headers = ["Α/Α", "ΟΝΟΜΑΤΕΠΩΝΥΜΟ", "ΓΕΝ.", "ΣΩΜΑΤΕΙΟ", "ΕΠΙΔ.", "ΑΝΕΜ.", "ΑΓΩΝΑΣ", "ΗΜ/ΝΙΑ", "ΤΟΠΟΘΕΣΙΑ", "ΣΕΙΡΑ", "ΔΙΑΔ.", "ΣΗΜΕΙΩΣΕΙΣ"]
    all_pdf_rows = wind_legal_ranking + wind_aided_ranking
    total = len(all_pdf_rows)

    pdf.set_font("DejaVu", "", 6)
    col_w = []
    col_w.append(max(pdf.get_string_width(str(i)) for i in range(1, total + 1)) + 2)
    col_w.append(max(pdf.get_string_width(r["name"]) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(str(r["birth_year"])) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(r["club"]) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(r["performance"]) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(fmt_wind(r["wind"])) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(fmt_comp(r)) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(r["date"]) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(fmt_loc(r)) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(rget(r, "heat")) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(rget(r, "lane")) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(fmt_note(r)) for r in all_pdf_rows) + 2)
    col_w[11] = max(col_w[11], 8)

    pdf.set_font("DejaVu", "B", 7)
    for ci, h in enumerate(headers):
        hw = pdf.get_string_width(h) + 2
        if hw > col_w[ci]:
            col_w[ci] = hw

    def pdf_header():
        pdf.set_font("DejaVu", "B", 9)
        pdf.cell(0, 5, "200 Μ ΚΟΡΑΣΙΔΩΝ (Κ18) 2026", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("DejaVu", "B", 7)
        for ci, h in enumerate(headers):
            pdf.cell(col_w[ci], 5, h, border=1, align="C")
        pdf.ln()

    def pdf_row(i, r):
        pdf.set_font("DejaVu", "", 6)
        vals = [
            str(i), r["name"], str(r["birth_year"]), r["club"],
            r["performance"], fmt_wind(r["wind"]), fmt_comp(r), r["date"],
            fmt_loc(r), rget(r, "heat"), rget(r, "lane"), fmt_note(r)
        ]
        for ci, v in enumerate(vals):
            pdf.cell(col_w[ci], 4, v, border=1, align="C" if ci == 0 else "L")
        pdf.ln()

    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    pdf_header()

    for i, r in enumerate(wind_legal_ranking, 1):
        if pdf.y > 270:
            pdf.add_page()
            pdf_header()
        pdf_row(i, r)

    if wind_aided_ranking:
        if pdf.y > 260:
            pdf.add_page()
            pdf_header()
        pdf.set_font("DejaVu", "B", 7)
        pdf.cell(sum(col_w), 5, "ΜΕ ΑΝΕΜΟ", border=1, align="C")
        pdf.ln()
        for i, r in enumerate(wind_aided_ranking, 1):
            if pdf.y > 270:
                pdf.add_page()
                pdf_header()
            pdf_row(i, r)

    pdf.output(pdf_filename)
    print(f"PDF file: {pdf_filename}")


def open_file(fpath):
    if sys.platform == "win32":
        os.startfile(fpath)
    elif sys.platform == "darwin":
        subprocess.run(["open", fpath])
    else:
        subprocess.run(["xdg-open", fpath])

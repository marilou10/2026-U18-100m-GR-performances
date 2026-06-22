# Copyright (c) 2026 Μαρια Ελένη Αντωνοπούλου
# Licensed under the MIT License. See LICENSE file.
#
# 2026 U18 100m Greek performances — scraper & Excel/PDF generator.
# Scrapes Roster Athletics links, normalizes names/clubs, produces
# All_Performances / Season_Best / Καλύτερες_Επιδόσεις sheets.

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from datetime import datetime

try:
    from fpdf import FPDF
    _HAS_PDF = True
except ImportError:
    _HAS_PDF = False
import os
import json
import time
import sys
import subprocess
import re
import unicodedata


LINKS_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "meet_links.txt"
)

CACHE_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "cache_performances.json"
)

# Load URLs and auto-convert about?id= to schedule?id=
fixed_links = []
with open(LINKS_FILE, 'r', encoding="utf-8") as f:
    raw_lines = f.readlines()

for line in raw_lines:
    stripped = line.strip()
    if not stripped:
        fixed_links.append(line)
        continue
    original = stripped
    stripped = stripped.replace("about?id=", "schedule?id=")
    fixed_links.append(stripped + "\n")

URLS = [l.strip() for l in fixed_links if l.strip()]

# Save back if any about?id= were replaced
if any("about?id=" in l for l in raw_lines):
    with open(LINKS_FILE, 'w', encoding="utf-8") as f:
        f.writelines(fixed_links)
    print("[OK] Auto-converted about?id= to schedule?id= in meet_links.txt")

all_results = []
scraped_urls = set()


def perf_float(p):
    if not p:
        return 999.0
    clean = re.sub(r'[^\d.]', '', p.replace("(", "").split()[0])
    # Handle empty or malformed (e.g. "..", ".") values
    try:
        return float(clean) if clean else 999.0
    except ValueError:
        return 999.0


GREEK_TO_LATIN = {
    ord('Α'): 'A', ord('α'): 'a',
    ord('Β'): 'V', ord('β'): 'v',
    ord('Γ'): 'G', ord('γ'): 'g',
    ord('Δ'): 'D', ord('δ'): 'd',
    ord('Ε'): 'E', ord('ε'): 'e',
    ord('Ζ'): 'Z', ord('ζ'): 'z',
    ord('Η'): 'I', ord('η'): 'i',
    ord('Θ'): 'TH', ord('θ'): 'th',
    ord('Ι'): 'I', ord('ι'): 'i',
    ord('Κ'): 'K', ord('κ'): 'k',
    ord('Λ'): 'L', ord('λ'): 'l',
    ord('Μ'): 'M', ord('μ'): 'm',
    ord('Ν'): 'N', ord('ν'): 'n',
    ord('Ξ'): 'X', ord('ξ'): 'x',
    ord('Ο'): 'O', ord('ο'): 'o',
    ord('Π'): 'P', ord('π'): 'p',
    ord('Ρ'): 'R', ord('ρ'): 'r',
    ord('Σ'): 'S', ord('σ'): 's', ord('ς'): 's',
    ord('Τ'): 'T', ord('τ'): 't',
    ord('Υ'): 'Y', ord('υ'): 'y',
    ord('Φ'): 'F', ord('φ'): 'f',
    ord('Χ'): 'CH', ord('χ'): 'ch',
    ord('Ψ'): 'PS', ord('ψ'): 'ps',
    ord('Ω'): 'O', ord('ω'): 'o',
    ord('Ά'): 'A', ord('ά'): 'a',
    ord('Έ'): 'E', ord('έ'): 'e',
    ord('Ή'): 'I', ord('ή'): 'i',
    ord('Ί'): 'I', ord('ί'): 'i', ord('ΐ'): 'i',
    ord('Ό'): 'O', ord('ό'): 'o',
    ord('Ύ'): 'Y', ord('ύ'): 'y', ord('ΰ'): 'y',
    ord('Ώ'): 'O', ord('ώ'): 'o',
}

def normalize_name(name):
    """Transliterate Greek to Latin so ΓΕΩΡΓΙΑ ΣΤΑΘΟΠΟΥΛΟΥ and GEORGIA STATHOPOULOU match."""
    return name.translate(GREEK_TO_LATIN).upper().strip()

def nk_latin(name):
    """Extract surname as Latin string for cross-script matching."""
    n = normalize_name(name)
    n = n.replace("OY", "OU")
    parts = n.split()
    return parts[-1] if parts else ""

def norm_full(name):
    """Full name normalized to Latin, for cross-script identity matching."""
    n = normalize_name(name)
    n = n.replace("OY", "OU").replace("GG", "NG")
    return n


# =========================
# LOAD EXISTING CACHE
# =========================
if os.path.exists(CACHE_FILE):
    print("Loading existing cache...")
    try:
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, dict) and "performances" in data:
            all_results = data["performances"]
            scraped_urls = set(data.get("scraped_urls", []))
        elif isinstance(data, list):
            all_results = data
        # Filter out entries with unparseable or non-100m performances
        before = len(all_results)
        all_results = [r for r in all_results if 0 < perf_float(r["performance"]) < 25.0]
        if len(all_results) < before:
            print(f"  Removed {before - len(all_results)} corrupted/non-100m entry(ies)")
        print(f"Loaded {len(all_results)} performances from cache\n")
    except (json.JSONDecodeError, KeyError):
        print("[WARN] Cache file is corrupted. Starting fresh...")
        all_results = []

# =========================
# DETERMINE NEW URLS
# =========================
urls_to_scrape = [url for url in URLS if url not in scraped_urls]
print(f"Total links in file: {len(URLS)} | Already cached: {len(URLS) - len(urls_to_scrape)} | New: {len(urls_to_scrape)}")

if urls_to_scrape:
    print(f"Scraping {len(urls_to_scrape)} new link(s)...\n")

    # =========================
    # CHROME OPTIONS
    # =========================
    chrome_options = Options()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
    chrome_options.add_argument("--disable-dev-shm-features")
    chrome_options.add_argument("--no-sandbox")

    # =========================
    # SCRAPE LOOP (PARALLEL)
    # =========================
    from concurrent.futures import ThreadPoolExecutor, as_completed

    NUM_WORKERS = 3

    PERFORMANCE_RE = re.compile(r"\d+\.\d+")

    def find_performance_cell(row):
        cells = row.find_elements(By.TAG_NAME, "td")
        for idx, cell in enumerate(cells):
            text = cell.text.split("\n")[0].strip()
            if "DNS" in text or "DNF" in text:
                continue
            m = PERFORMANCE_RE.search(text)
            if m:
                perf = m.group(0)
                # Re-append trailing suffix chars (w, h) from original text
                rest = text[m.end():]
                for suffix in ("w", "h"):
                    if suffix in rest:
                        perf += suffix
                        break
                return perf, idx
        return None, None

    def start_driver():
        return webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=chrome_options
        )

    def worker(urls_batch, worker_id):
        driver = start_driver()
        local_results = []
        succeeded = []
        try:
            for url_index, url in enumerate(urls_batch):
                print(f"\n[W{worker_id}] ({url_index + 1}/{len(urls_batch)}) {url}")
                try:
                    driver.get(url)

                    try:
                        WebDriverWait(driver, 15).until(
                            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tbody tr"))
                        )
                    except TimeoutException:
                        print(f"[W{worker_id}]   [WARN] Timeout: Could not load schedule from {url}")
                        succeeded.append(url)
                        continue

                    time.sleep(1)

                    # =========================
                    # SCRAPE MEET INFO
                    # =========================
                    body_text = driver.find_element(By.TAG_NAME, "body").text
                    lines = body_text.split("\n")

                    meet_name = ""
                    meet_date = ""
                    meet_location = ""

                    for i, line in enumerate(lines):
                        line = line.strip()
                        if line == "ΟΝΟΜΑ" and i + 1 < len(lines):
                            raw = lines[i + 1].strip()
                            if " · " in raw:
                                meet_name = raw.split(" · ", 1)[1].strip()
                            else:
                                meet_name = raw
                        if line == "ΗΜΕΡΟΜΗΝΙΑ & ΩΡΑ" and i + 1 < len(lines):
                            raw = lines[i + 1].strip()
                            meet_date = raw[:10]
                        if line == "ΠΟΛΗ & ΧΩΡΑ" and i + 1 < len(lines):
                            raw = lines[i + 1].strip()
                            meet_location = re.sub(r'^\d+\s*', '', raw).strip()

                    meet_name = re.sub(r'^Roster Athletics\s*·\s*', '', meet_name).strip()

                    if not meet_name:
                        meet_name = driver.title.strip()

                    print(f"[W{worker_id}]   Meet: {meet_name} | Date: {meet_date} | Location: {meet_location}")

                    # =========================
                    # FIND WOMEN'S 100m FINALS
                    # =========================
                    rows = driver.find_elements(By.CSS_SELECTOR, "tbody tr")
                    print(f"[W{worker_id}]   Found {len(rows)} rows in schedule")

                    matching_indices = []
                    for i, row in enumerate(rows):
                        text = row.text.strip()
                        if (
                            "100μ" in text
                            and "Τελικός" in text
                            and "Εμπόδια" not in text
                            and "4Χ100" not in text.upper()
                            and "4X100" not in text.upper()
                            and ("Γυναίκες" in text or "Κορίτσια" in text)
                        ):
                            matching_indices.append(i)
                            print(f"[W{worker_id}]   [OK] Matched row {i}: {repr(text)}")

                    print(f"[W{worker_id}]   Found {len(matching_indices)} matching women's 100m finals rows")

                    results_urls = []

                    for i in matching_indices:
                        try:
                            WebDriverWait(driver, 15).until(
                                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tbody tr"))
                            )
                            time.sleep(0.5)
                            rows = driver.find_elements(By.CSS_SELECTOR, "tbody tr")

                            if i >= len(rows):
                                print(f"[W{worker_id}]   [WARN] Row {i} no longer exists (only {len(rows)} rows)")
                                continue

                            row = rows[i]
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row)
                            time.sleep(0.5)

                            anchors = row.find_elements(By.TAG_NAME, "a")
                            if anchors:
                                href = anchors[0].get_attribute("href")
                                if href:
                                    results_urls.append(href)
                                    print(f"[W{worker_id}]   [OK] Found results URL: {href}")
                                    continue

                            td = row.find_elements(By.TAG_NAME, "td")
                            target = td[1] if len(td) > 1 else row

                            actions = ActionChains(driver)
                            actions.move_to_element(target).click().perform()

                            try:
                                WebDriverWait(driver, 10).until(
                                    lambda d: "meId" in d.current_url
                                )
                            except TimeoutException:
                                print(f"[W{worker_id}]   [WARN] URL did not change after clicking row {i}")
                                driver.back()
                                WebDriverWait(driver, 15).until(
                                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tbody tr"))
                                )
                                continue

                            current = driver.current_url
                            results_urls.append(current)
                            print(f"[W{worker_id}]   [OK] Found results URL: {current}")
                            driver.back()
                            WebDriverWait(driver, 15).until(
                                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tbody tr"))
                            )

                        except Exception as click_error:
                            print(f"[W{worker_id}]   [WARN] Could not click row {i}: {click_error}")
                            continue

                    if not results_urls:
                        print(f"[W{worker_id}]   [WARN] No women's 100m finals found in {url}")
                        succeeded.append(url)
                        continue

                    # =========================
                    # SCRAPE EACH RESULTS PAGE
                    # =========================
                    for results_url in results_urls:
                        print(f"[W{worker_id}]   Scraping results: {results_url}")

                        try:
                            driver.get(results_url)

                            try:
                                WebDriverWait(driver, 15).until(
                                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tbody tr"))
                                )
                            except TimeoutException:
                                print(f"[W{worker_id}]   [WARN] Timeout loading results: {results_url}")
                                continue

                            time.sleep(1)

                            body_text = driver.find_element(By.TAG_NAME, "body").text
                            lines = body_text.split("\n")

                            heat_name = ""
                            has_100m = False
                            for line in lines:
                                line = line.strip()
                                if "100μ" in line and "Τελικός" in line:
                                    has_100m = True
                                    parts = line.split("Τελικός")
                                    if len(parts) > 1:
                                        after = parts[1].strip()
                                        letter = after.split()[0] if after else ""
                                        heat_name = f"Τελικός {letter}" if letter else "Τελικός"
                                    break

                            if not has_100m:
                                print(f"[W{worker_id}]   [SKIP] No 100m event on this results page")
                                continue

                            wind = ""
                            for line in lines:
                                line = line.strip()
                                if line.startswith("Άνεμος:"):
                                    wind = line.replace("Άνεμος:", "").strip()
                                    break

                            # Find the specific table belonging to the 100m event (skip other tables on the page)
                            try:
                                tables = driver.find_elements(By.TAG_NAME, "table")
                                table_100m = None
                                for t in tables:
                                    before = t.find_elements(By.XPATH, "./preceding::*[contains(., '100μ') and contains(., 'Τελικός')][1]")
                                    if before:
                                        table_100m = t
                                        break
                                if table_100m:
                                    result_rows = table_100m.find_elements(By.CSS_SELECTOR, "tbody tr")
                                else:
                                    result_rows = driver.find_elements(By.CSS_SELECTOR, "tbody tr")
                            except Exception:
                                print(f"[W{worker_id}]   [WARN] Could not isolate 100m table, using all rows")
                                result_rows = driver.find_elements(By.CSS_SELECTOR, "tbody tr")
                            print(f"[W{worker_id}]   [OK] Βρέθηκαν {len(result_rows)} γραμμές")

                            for row in result_rows:
                                try:
                                    cells = row.find_elements(By.TAG_NAME, "td")

                                    if len(cells) < 5:
                                        continue

                                    lane = cells[1].text.strip()
                                    athlete_info = cells[2].text.split("\n")
                                    name = athlete_info[0].strip()

                                    if len(athlete_info) < 2:
                                        continue

                                    try:
                                        birth_year = int(athlete_info[1].strip())
                                    except ValueError:
                                        continue

                                    performance, perf_idx = find_performance_cell(row)

                                    if not performance:
                                        continue

                                    if perf_float(performance) > 25.0:
                                        continue

                                    all_cells = row.find_elements(By.TAG_NAME, "td")
                                    club = all_cells[perf_idx - 1].text.strip() if perf_idx and perf_idx > 0 else ""
                                    if club and name and club == name.split()[-1]:
                                        club = ""

                                    if 2009 <= birth_year <= 2012:
                                        local_results.append({
                                            "name": name,
                                            "birth_year": birth_year,
                                            "club": club,
                                            "performance": performance,
                                            "wind": wind,
                                            "competition": meet_name,
                                            "date": meet_date,
                                            "location": meet_location,
                                            "heat": heat_name,
                                            "lane": lane
                                        })
                                except Exception as row_error:
                                    print(f"[W{worker_id}]     [WARN] Error parsing row: {row_error}")
                                    continue

                        except Exception as results_error:
                            print(f"[W{worker_id}]   [ERR] Error scraping results page: {results_error}")
                            continue

                except Exception as e:
                    print(f"[W{worker_id}]   [ERR] Error scraping {url}: {e}")
                    if "invalid session id" in str(e).lower() or "no such window" in str(e).lower():
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        driver = start_driver()
                    continue

                succeeded.append(url)
        finally:
            try:
                driver.quit()
            except Exception:
                pass
        return local_results, succeeded

    # Distribute URLs evenly across workers
    batches = [[] for _ in range(NUM_WORKERS)]
    for i, url in enumerate(urls_to_scrape):
        batches[i % NUM_WORKERS].append(url)
    batches = [b for b in batches if b]

    all_new_results = []
    all_succeeded = []
    with ThreadPoolExecutor(max_workers=len(batches)) as executor:
        futures = {executor.submit(worker, batch, i): i for i, batch in enumerate(batches)}
        for future in as_completed(futures):
            try:
                results, succeeded = future.result()
                all_new_results.extend(results)
                all_succeeded.extend(succeeded)
            except Exception as e:
                print(f"[ERR] Worker failed: {e}")

    all_results.extend(all_new_results)
    scraped_urls.update(all_succeeded)

    # Remove duplicates
    unique = {}
    for r in all_results:
        key = (
            r["name"],
            r["birth_year"],
            r["competition"],
            r["performance"]
        )
        unique[key] = r

    all_results = list(unique.values())

# =========================
# CLEAN COMPETITION NAMES & LOCATIONS
# =========================
for r in all_results:
    r["competition"] = re.sub(r'^Roster Athletics\s*·\s*', '', r["competition"]).strip()
    r["location"] = re.sub(r'^\d+\s*', '', r["location"]).strip()
    r["location"] = re.sub(r', ([^,]+), \1$', r', \1', r["location"])
    m = re.search(r'& (.+?) \(.*?(?:ΔΡ[ΟΌ]ΜΟΙ|[ΑΆ]ΛΜΑΤΑ).*?\), (.+)', r["location"], re.IGNORECASE)
    if m:
        r["location"] = m.group(1).strip() + ", " + m.group(2).strip()
    # Strip embedded location from competition name
    loc_city = r["location"].split(",")[0].strip()
    if loc_city:
        comp = r["competition"]
        def _no_tonos(s):
            return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn').upper()
        comp_flat = _no_tonos(comp)
        city_flat = _no_tonos(loc_city)
        idx = comp.find(",")
        if idx >= 0 and city_flat in comp_flat[idx:idx+50]:
            r["competition"] = comp[:idx].strip()

# =========================
# TRANSLATE LATIN LOCATIONS TO GREEK
# =========================
LOCATION_GR = {
    "ALEXANDRIA": "ΑΛΕΞΑΝΔΡΕΙΑ",
    "ARGOSTOLI, KEFALONIA": "ΑΡΓΟΣΤΟΛΙ, ΚΕΦΑΛΟΝΙΑ",
    "ATHENS": "ΑΘΗΝΑ",
    "NAXOS KAI MIKRES KYKLADES": "ΝΑΞΟΣ ΚΑΙ ΜΙΚΡΕΣ ΚΥΚΛΑΔΕΣ",
    "TRIKALA": "ΤΡΙΚΑΛΑ",
    "VARI ATHINA": "ΒΑΡΗ ΑΘΗΝΑ",
    "VARI ATHINA, GREECE": "ΒΑΡΗ ΑΘΗΝΑ",
    "ALEXANDRIA, GREECE": "ΑΛΕΞΑΝΔΡΕΙΑ",
    "ARGOSTOLI, KEFALONIA, GREECE": "ΑΡΓΟΣΤΟΛΙ, ΚΕΦΑΛΟΝΙΑ",
    "ATHENS, GREECE": "ΑΘΗΝΑ",
    "NAXOS KAI MIKRES KYKLADES, GREECE": "ΝΑΞΟΣ ΚΑΙ ΜΙΚΡΕΣ ΚΥΚΛΑΔΕΣ",
    "TRIKALA, GREECE": "ΤΡΙΚΑΛΑ",
}
for r in all_results:
    loc_up = r["location"].upper()
    if loc_up in LOCATION_GR:
        r["location"] = LOCATION_GR[loc_up]

# Backfill missing clubs from other entries of the same athlete (cross-script)
# Treat "Greece" placeholder club as missing
PLACEHOLDER_CLUBS = {"GREECE", "Greece"}
def _club_key(n):
    """Standard + alternative V↔Y key for cross-script matching (e.g. V↔Y from Y)."""
    k = norm_full(n)
    return k, k.replace("V", "Y").replace("B", "Y")

club_lookup = {}
for r in all_results:
    c = r.get("club", "").strip()
    if c and c not in PLACEHOLDER_CLUBS:
        k1, k2 = _club_key(r["name"])
        club_lookup.setdefault(k1, set()).add(c)
        club_lookup.setdefault(k2, set()).add(c)
for r in all_results:
    c = r.get("club", "").strip()
    if not c or c in PLACEHOLDER_CLUBS:
        k1, k2 = _club_key(r["name"])
        for k in (k1, k2):
            if k in club_lookup:
                r["club"] = list(club_lookup[k])[0]
                break
    # Clear remaining placeholders
    if r.get("club", "").strip() in PLACEHOLDER_CLUBS:
        r["club"] = ""

# Clean malformed performances (e.g. "13.65 (.641)" -> "13.65")
PERF_CLEAN_RE = re.compile(r"(\d+\.\d+)")
for r in all_results:
    p = r["performance"]
    if "(" in p or ")" in p:
        m = PERF_CLEAN_RE.search(p)
        if m:
            clean = m.group(1)
            if "w" in p:
                clean += "w"
            if "h" in p:
                clean += "h"
            r["performance"] = clean

# Known over-age athletes mistakenly included (Roster birth year incorrect)
OVERRIDE_EXCLUDE = {
    "ΚΑΛΛΙΟΠΙ ΠΑΥΛΑΚΑΚΗ",  # born 2002, not U18
    "Kalliopi PAVLAKAKI",   # born 2002, not U18
}
all_results = [r for r in all_results if r["name"] not in OVERRIDE_EXCLUDE]

# Non-Greek clubs (athletes from Cyprus, Bulgaria, etc.)
NON_GREEK_CLUBS = {
    "ΑΠΟΕΛ",
    "ΛΥΚΕΙΟ ΒΕΡΓΙΝΑΣ - ΚΥΠΡΟΣ",
    "KLASA",
    "SKLA Atlet - Mezdra",
    'ASC "Lokomotiv - Ruse"',
    "Priority Sport",
    "Sundsvalls FI",
}
before = len(all_results)
all_results = [r for r in all_results if r.get("club", "").strip() not in NON_GREEK_CLUBS]
if len(all_results) < before:
    print(f"[OK] Removed {before - len(all_results)} entries by non-Greek club athletes")

if urls_to_scrape:
    # =========================
    # SAVE CACHE
    # =========================
    print(f"\nSaving {len(all_results)} performances to cache...")
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump({
            "performances": all_results,
            "scraped_urls": sorted(scraped_urls)
        }, f, ensure_ascii=False, indent=2)
    print("[OK] Cache saved successfully")
else:
    print("All links already scraped. Using cached data.")
    print(f"Saving {len(all_results)} performances to cache...")
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump({
            "performances": all_results,
            "scraped_urls": sorted(scraped_urls)
        }, f, ensure_ascii=False, indent=2)
    print("[OK] Cache saved successfully")

# =========================
# NORMALIZE NAMES TO GREEK
# =========================
# Build a mapping of (normalized_name, year) -> preferred Greek name
preferred_names = {}
for r in all_results:
    has_greek = any('\u0370' <= c <= '\u03FF' for c in r['name'])
    key = (norm_full(r["name"]), r["birth_year"])
    if has_greek and key not in preferred_names:
        preferred_names[key] = r['name']

# Replace Latin names with their Greek equivalent where available
for r in all_results:
    has_greek = any('\u0370' <= c <= '\u03FF' for c in r['name'])
    if not has_greek:
        key = (norm_full(r["name"]), r["birth_year"])
        if key in preferred_names:
            r['name'] = preferred_names[key]

# Second pass: surname-only fallback for single-match cases (e.g. "Georgia Rafailia KANLI")
surname_greek = {}
for r in all_results:
    has_greek = any('\u0370' <= c <= '\u03FF' for c in r['name'])
    if has_greek:
        key = (nk_latin(r['name']), r['birth_year'])
        if key not in surname_greek:
            surname_greek[key] = set()
        surname_greek[key].add(r['name'])

for r in all_results:
    has_greek = any('\u0370' <= c <= '\u03FF' for c in r['name'])
    if not has_greek:
        key = (nk_latin(r['name']), r['birth_year'])
        if key in surname_greek and len(surname_greek[key]) == 1:
            r['name'] = list(surname_greek[key])[0]

# Third pass: Latin->Greek reverse transliteration for still-unmatched Greek names
LATIN_TO_GREEK_DIGRAPHS = [
    ("CH", "Χ"), ("TH", "Θ"), ("PS", "Ψ"), ("OU", "ΟΥ"),
    ("MP", "ΜΠ"), ("NT", "ΝΤ"), ("GK", "ΓΚ"), ("NG", "ΓΓ"),
    ("TS", "ΤΣ"), ("TZ", "ΤΖ"), ("AI", "ΑΙ"), ("EI", "ΕΙ"),
    ("OI", "ΟΙ"), ("AY", "ΑΥ"), ("EY", "ΕΥ"),
    ("AV", "ΑΥ"), ("EV", "ΕΥ"),  # V as second in diphthong -> Υ not Β
]
LATIN_TO_GREEK_SINGLE = {
    'A': 'Α', 'B': 'Β', 'C': 'Σ', 'D': 'Δ', 'E': 'Ε',
    'F': 'Φ', 'G': 'Γ', 'I': 'Ι', 'K': 'Κ', 'L': 'Λ',
    'M': 'Μ', 'N': 'Ν', 'O': 'Ο', 'P': 'Π', 'R': 'Ρ',
    'S': 'Σ', 'T': 'Τ', 'U': 'ΟΥ', 'V': 'Β', 'X': 'Ξ',
    'Y': 'Υ', 'Z': 'Ζ',
}

def latin_to_greek(name):
    result = []
    i = 0
    upper = name.upper()
    while i < len(upper):
        matched = False
        for digraph, greek in LATIN_TO_GREEK_DIGRAPHS:
            if upper[i:i+len(digraph)] == digraph:
                result.append(greek)
                i += len(digraph)
                matched = True
                break
        if matched:
            continue
        ch = upper[i]
        result.append(LATIN_TO_GREEK_SINGLE.get(ch, ch))
        i += 1
    out = "".join(result)
    # Common Greek surname ending correction: -ΚΙ -> -ΚΗ
    parts = out.split()
    if parts and len(parts[-1]) > 2 and parts[-1].endswith("ΚΙ"):
        parts[-1] = parts[-1][:-1] + "Η"
        out = " ".join(parts)
    return out

for r in all_results:
    has_greek = any('\u0370' <= c <= '\u03FF' for c in r['name'])
    if not has_greek:
        greek_attempt = latin_to_greek(r['name'])
        # Only use if all characters successfully mapped (no leftover ASCII letters)
        if not any('A' <= c <= 'Z' for c in greek_attempt.upper()):
            # Only convert if name contains at least one strong Greek digraph
            # (avoids false conversion of non-Greek names like NIKOLOVA, DAHLGREN)
            GREEK_STRONG_DIGRAPHS = {"CH","TH","PS","OU","MP","NT","GK","NG","TZ","AY","EY","AV","EV"}
            upper_name = r['name'].upper()
            surname = upper_name.split()[-1] if ' ' in upper_name else upper_name
            has_strong = False
            for d in GREEK_STRONG_DIGRAPHS:
                if d not in upper_name:
                    continue
                # Skip "EV" when it's part of "-EVA" Bulgarian patronymic suffix
                if d == "EV" and surname.endswith("EVA") and "EV" in surname[-3:]:
                    continue
                has_strong = True
                break
            if has_strong:
                r['name'] = greek_attempt

# =========================
# FILTER NON-GREEK ATHLETES
# =========================
# Athletes whose names remain in Latin after all normalization passes are non-Greek
before = len(all_results)
all_results = [r for r in all_results if any('\u0370' <= c <= '\u03FF' for c in r['name'])]
removed = before - len(all_results)
if removed:
    print(f"[OK] Removed {removed} entries by non-Greek athletes")

# Manual club fixups for athletes where backfill failed (transliteration mismatch)
# Match on normalized surname + first-name initial to catch spelling variants
MANUAL_CLUB = {
    "KANLI": "ΓΑΣ ΜΗΘΥΜΝΑΣ ΟΛΥΜΠΙΑΣ ΛΕΣ",
}
for r in all_results:
    if r.get("club", "").strip():
        continue
    sur = nk_latin(r["name"])
    if sur in MANUAL_CLUB:
        r["club"] = MANUAL_CLUB[sur]

# =========================
# SEASON BEST
# =========================
season_best = {}

for r in all_results:
    try:
        perf = perf_float(r["performance"])
    except ValueError:
        continue

    key = (norm_full(r["name"]), r["birth_year"])

    if key not in season_best:
        season_best[key] = r
    else:
        if perf < perf_float(season_best[key]["performance"]):
            season_best[key] = r

ranking = sorted(
    season_best.values(),
    key=lambda x: perf_float(x["performance"])
)

# =========================
# WIND-LEGAL BEST
# =========================
def is_wind_legal(r):
    if 'w' in r['performance']:
        return False
    w = r['wind'].strip().lstrip('+')
    if w.upper() == 'NWI':
        return True
    try:
        return float(w) <= 2.0
    except ValueError:
        return True

wind_legal_best = {}
for r in all_results:
    if not is_wind_legal(r):
        continue
    try:
        perf = perf_float(r["performance"])
    except ValueError:
        continue
    key = (norm_full(r["name"]), r["birth_year"])
    if key not in wind_legal_best or perf < perf_float(wind_legal_best[key]["performance"]):
        wind_legal_best[key] = r

wind_legal_ranking = sorted(
    wind_legal_best.values(),
    key=lambda x: perf_float(x["performance"])
)

# =========================
# WIND-AIDED ONLY (athletes with no wind-legal performances at all)
# =========================
wind_aided_only = {}
for r in all_results:
    if is_wind_legal(r):
        continue
    try:
        perf = perf_float(r["performance"])
    except ValueError:
        continue
    key = (norm_full(r["name"]), r["birth_year"])
    # Only include if athlete has NO wind-legal entry
    if key in wind_legal_best:
        continue
    if key not in wind_aided_only or perf < perf_float(wind_aided_only[key]["performance"]):
        wind_aided_only[key] = r

wind_aided_ranking = sorted(
    wind_aided_only.values(),
    key=lambda x: perf_float(x["performance"])
)

# =========================
# UPPERCASE ALL TEXT FIELDS FOR OUTPUT
# =========================
TEXT_FIELDS = ["name", "club", "competition", "location", "date", "heat", "lane"]
for r in all_results:
    for k in TEXT_FIELDS:
        v = r.get(k)
        if v and isinstance(v, str):
            r[k] = v.upper()

# Clear any remaining placeholder clubs after uppercase
for r in all_results:
    if r.get("club", "").strip() in PLACEHOLDER_CLUBS:
        r["club"] = ""

# Clean up placeholder lane values
for r in all_results:
    lane = r.get("lane", "").strip()
    if lane in ("", "-", "--"):
        r["lane"] = ""

# =========================
# LOAD NOTES
# =========================
NOTES_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "cache_notes.json"
)
_notes_data = {}
if os.path.exists(NOTES_FILE):
    try:
        with open(NOTES_FILE, 'r', encoding='utf-8') as f:
            _notes_data = json.load(f)
    except (json.JSONDecodeError, IOError):
        _notes_data = {}

def fmt_note(r):
    k = f"{norm_full(r['name'])}|{r.get('birth_year','')}|{r.get('date','')}"
    return _notes_data.get(k, "")

# =========================
# OUTPUT FILE
# =========================
base_dir = os.path.dirname(os.path.abspath(__file__))
output_dir = os.path.join(base_dir, "output")
os.makedirs(output_dir, exist_ok=True)

timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")

filename = os.path.join(
    output_dir,
    f"2026_U18_100m_GR_{timestamp}.xlsx"
)

def fmt_wind(w):
    w = w.strip()
    if not w:
        return "NWI"
    if w.upper() == "NWI":
        return "NWI"
    return w.lstrip("+")

def fmt_comp(r):
    return r["competition"]

def fmt_loc(r):
    loc = r["location"]
    if loc.endswith(", GREECE"):
        loc = loc[:-8]
    return loc

G = ["Α/Α","ΟΝΟΜΑΤΕΠΩΝΥΜΟ","ΓΕΝΝΗΣΗ","ΣΩΜΑΤΕΙΟ","ΕΠΙΔΟΣΗ","ΑΝΕΜΟΣ","ΑΓΩΝΑΣ","ΗΜ/ΝΙΑ","ΤΟΠΟΘΕΣΙΑ","ΣΕΙΡΑ","ΔΙΑΔΡΟΜΟΣ","ΣΗΜΕΙΩΣΕΙΣ"]

wb = Workbook()

ws1 = wb.active
ws1.title = "All_Performances"
ws1.append(G)

sorted_all = sorted(all_results, key=lambda x: perf_float(x["performance"]))
for i, r in enumerate(sorted_all, 1):
    ws1.append([i, r["name"], r["birth_year"], r["club"], r["performance"], fmt_wind(r["wind"]), fmt_comp(r), r["date"], fmt_loc(r), r["heat"], r["lane"], fmt_note(r)])

ws2 = wb.create_sheet("Season_Best")
ws2.append(["100 Μ ΚΟΡΑΣΙΔΩΝ (Κ18) 2026"])
ws2.append(G)
for i, r in enumerate(ranking, 1):
    ws2.append([i, r["name"], r["birth_year"], r["club"], r["performance"], fmt_wind(r["wind"]), fmt_comp(r), r["date"], fmt_loc(r), r["heat"], r["lane"], fmt_note(r)])

ws3 = wb.create_sheet("Καλύτερες_Επιδόσεις")
ws3.append(["100 Μ ΚΟΡΑΣΙΔΩΝ (Κ18) 2026"])
ws3.append(G)

for i, r in enumerate(wind_legal_ranking, 1):
        ws3.append([i, r["name"], r["birth_year"], r["club"], r["performance"], fmt_wind(r["wind"]), fmt_comp(r), r["date"], fmt_loc(r), r["heat"], r["lane"], fmt_note(r)])

if wind_aided_ranking:
    ws3.append([])
    ws3.append(["ΜΕ ΑΝΕΜΟ"] + [""] * 11)
    for i, r in enumerate(wind_aided_ranking, 1):
        ws3.append([i, r["name"], r["birth_year"], r["club"], r["performance"], fmt_wind(r["wind"]), fmt_comp(r), r["date"], fmt_loc(r), r["heat"], r["lane"], fmt_note(r)])

COPYRIGHT = "Copyright (c) 2026 Μαρια Ελένη Αντωνοπούλου — Licensed under MIT"
ws1.append([])
ws1.append([COPYRIGHT] + [""] * 11)
ws2.append([])
ws2.append([COPYRIGHT] + [""] * 11)
ws3.append([])
ws3.append([COPYRIGHT] + [""] * 11)

wb.save(filename)

# =========================
# PDF EXPORT
# =========================
if _HAS_PDF:
    pdf_filename = filename.replace(".xlsx", ".pdf")

    class PDF(FPDF):
        def footer(self):
            self.set_y(-10)
            self.set_font("DejaVu", "", 5)
            self.cell(0, 5, COPYRIGHT, align="C")
    pdf = PDF(orientation="L", unit="mm", format="A3")
    pdf.set_left_margin(5)

    pdf.add_font("DejaVu", "", r"C:\Windows\Fonts\DejaVuSans.ttf")
    pdf.add_font("DejaVu", "B", r"C:\Windows\Fonts\DejaVuSans-Bold.ttf")

    headers = ["Α/Α", "ΟΝΟΜΑΤΕΠΩΝΥΜΟ", "ΓΕΝ.", "ΣΩΜΑΤΕΙΟ", "ΕΠΙΔ.", "ΑΝΕΜ.", "ΑΓΩΝΑΣ", "ΗΜ/ΝΙΑ", "ΤΟΠΟΘΕΣΙΑ", "ΣΕΙΡΑ", "ΔΙΑΔ.", "ΣΗΜ."]
    all_pdf_rows = wind_legal_ranking + wind_aided_ranking
    total = len(all_pdf_rows)
    total_aided = len(wind_aided_ranking)

    def rget(r, k):
        v = r.get(k)
        return str(v) if v is not None else ""

    pdf.set_font("DejaVu", "", 6)
    col_w = []
    col_w.append(max(pdf.get_string_width(str(i)) for i in range(1, total + 1)) + 2)
    col_w.append(max(pdf.get_string_width(r["name"]) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(str(r["birth_year"])) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(r["club"]) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(r["performance"]) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(fmt_wind(r["wind"])) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(fmt_comp(r)) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(r["date"]) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(fmt_loc(r)) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(rget(r, "heat")) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(rget(r, "lane")) for r in all_pdf_rows) + 2)
    col_w.append(max(pdf.get_string_width(fmt_note(r)) for r in all_pdf_rows) + 2)
    col_w[11] = max(col_w[11], 8)  # minimum 8mm

    pdf.set_font("DejaVu", "B", 7)
    for ci, h in enumerate(headers):
        hw = pdf.get_string_width(h) + 2
        if hw > col_w[ci]:
            col_w[ci] = hw

    def pdf_header():
        pdf.set_font("DejaVu", "B", 9)
        pdf.cell(0, 5, "100 Μ ΚΟΡΑΣΙΔΩΝ (Κ18) 2026", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("DejaVu", "B", 7)
        for ci, h in enumerate(headers):
            pdf.cell(col_w[ci], 5, h, border=1, align="C")
        pdf.ln()

    def pdf_row(i, r, bold=False):
        style = "B" if bold else ""
        pdf.set_font("DejaVu", style, 6)
        vals = [
            str(i), r["name"], str(r["birth_year"]), r["club"],
            r["performance"], fmt_wind(r["wind"]), fmt_comp(r), r["date"],
            fmt_loc(r), rget(r, "heat"), rget(r, "lane"), fmt_note(r)
        ]
        for ci, v in enumerate(vals):
            pdf.cell(col_w[ci], 4, v, border=1, align="C" if ci == 0 else "L")
        pdf.ln()

    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    pdf_header()

    for i, r in enumerate(wind_legal_ranking, 1):
        if pdf.y > 270:
            pdf.add_page()
            pdf_header()
        pdf_row(i, r)

    if wind_aided_ranking:
        if pdf.y > 260:
            pdf.add_page()
            pdf_header()
        pdf.set_font("DejaVu", "B", 7)
        pdf.cell(sum(col_w), 5, "ΜΕ ΑΝΕΜΟ", border=1, align="C")
        pdf.ln()
        for i, r in enumerate(wind_aided_ranking, 1):
            if pdf.y > 270:
                pdf.add_page()
                pdf_header()
            pdf_row(i, r)

    pdf.output(pdf_filename)
    print(f"PDF file: {pdf_filename}")
print("\n" + "="*50)
print("DONE")
print("="*50)
print(f"Excel file: {filename}")
print(f"Total performances: {len(all_results)}")
print(f"Season best athletes: {len(ranking)}")
print(f"Wind-legal best athletes: {len(wind_legal_ranking)}")

def _open(fpath):
    if sys.platform == "win32":
        os.startfile(fpath)
    elif sys.platform == "darwin":
        subprocess.run(["open", fpath])
    else:
        subprocess.run(["xdg-open", fpath])

_open(filename)
if _HAS_PDF:
    _open(pdf_filename)